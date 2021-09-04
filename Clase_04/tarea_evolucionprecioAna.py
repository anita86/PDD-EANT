# -*- coding: utf-8 -*-
"""
Created on Sun Jun 27 18:55:17 2021

@author: Ana
"""
from openpyxl import load_workbook

wb = load_workbook('lista_de_precios.xlsx', data_only=True)

sheets = wb.sheetnames
valores_precio  = []      
        
#Obtengo valor de la celda
for hoja in sheets: 
    hoja_activa = wb[hoja]
    precio = hoja_activa.cell(10, 6).value
    valores_precio.append(round(precio,2))
print(valores_precio)


  

    